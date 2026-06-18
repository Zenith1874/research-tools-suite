import io
import json
import re
import sqlite3
from datetime import datetime
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


PBOC_YEAR_INDEX = 'https://www.pbc.gov.cn/diaochatongjisi/116219/116319/index.html'
PBOC_SOURCE_NAME = '中国人民银行'
PBOC_BALANCE_SOURCE_TYPE = 'pboc_balance_sheet'

BASE_INDICATORS = {
    'total_assets': '资产总计',
    'foreign_assets': '国外资产合计',
    'foreign_exchange': '外汇',
    'monetary_gold': '货币黄金',
    'other_foreign_assets': '其他国外资产',
    'claims_on_government': '对政府债权',
    'claims_on_other_depository_corporations': '对其他存款性公司债权',
}

DERIVED_INDICATORS = {
    'foreign_assets_pct': ('国外资产 / 总资产', 'foreign_assets', 'total_assets'),
    'foreign_exchange_pct': ('外汇 / 总资产', 'foreign_exchange', 'total_assets'),
    'monetary_gold_pct': ('货币黄金 / 总资产', 'monetary_gold', 'total_assets'),
    'claims_on_government_pct': ('对政府债权 / 总资产', 'claims_on_government', 'total_assets'),
    'claims_on_other_depository_corporations_pct': (
        '对其他存款性公司债权 / 总资产',
        'claims_on_other_depository_corporations',
        'total_assets',
    ),
}

ROW_MATCHERS = [
    ('other_foreign_assets', ('其他国外资产', 'Other Foreign Assets')),
    ('foreign_exchange', ('外汇', 'Foreign Exchange')),
    ('monetary_gold', ('货币黄金', 'Monetary Gold')),
    ('foreign_assets', ('国外资产', 'Foreign Assets')),
    ('claims_on_other_depository_corporations', (
        '对其他存款性公司债权',
        'Claims on Other Depository Corporations',
    )),
    ('claims_on_government', ('对政府债权', 'Claims on Government')),
    ('total_assets', ('总资产', 'Total Assets')),
]


def connect(db_path):
    conn = sqlite3.connect(db_path, timeout=20)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_pboc_balance_sheet_tables(conn):
    conn.execute('''CREATE TABLE IF NOT EXISTS pboc_balance_sheet_observations (
        id INTEGER PRIMARY KEY,
        indicator_code TEXT,
        indicator_name TEXT,
        period TEXT,
        frequency TEXT,
        value REAL,
        unit TEXT,
        data_status TEXT,
        source_name TEXT,
        source_type TEXT,
        source_url TEXT,
        source_title TEXT,
        published_date TEXT,
        parser_notes TEXT,
        formula TEXT,
        updated_at TEXT,
        UNIQUE(indicator_code, period, source_url)
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS fiscal_debt_sources (
        id INTEGER PRIMARY KEY,
        source_name TEXT,
        source_type TEXT,
        source_url TEXT UNIQUE,
        source_title TEXT,
        published_date TEXT,
        parser_notes TEXT,
        raw_text TEXT,
        parsed_indicators TEXT,
        status TEXT,
        error TEXT,
        updated_at TEXT
    )''')


def fetch(url):
    r = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=30)
    r.raise_for_status()
    return r.content, r.url


def decode_html(content):
    return content.decode('utf-8', errors='replace')


def source_date_from_url(url):
    m = re.search(r'/(\d{4})/(\d{2})/(\d{8})', url)
    if m:
        d = m.group(3)
        return f'{d[:4]}-{d[4:6]}-{d[6:8]}'
    return None


def period_from_header(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        year = int(value)
        month = int(round((float(value) - year) * 100))
        if 1 <= month <= 12:
            return f'{year}-{month:02d}'
    s = str(value).strip().replace('年', '.').replace('月', '')
    m = re.search(r'(20\d{2})[.\-/](\d{1,2})', s)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}'
    return None


def normalize_label(value):
    return re.sub(r'\s+', ' ', str(value or '').replace('\xa0', ' ')).strip()


def discover_balance_sheet_attachment():
    content, final_url = fetch(PBOC_YEAR_INDEX)
    soup = BeautifulSoup(decode_html(content), 'html.parser')
    year_links = []
    for a in soup.find_all('a'):
        text = a.get_text(' ', strip=True)
        href = a.get('href')
        if href and re.search(r'20\d{2}年统计数据', text):
            year_links.append(urljoin(final_url, href))
    if not year_links:
        year_links = ['https://www.pbc.gov.cn/diaochatongjisi/116219/116319/2026ntjsj/index.html']

    errors = []
    for year_url in year_links[:4]:
        try:
            year_content, year_final = fetch(year_url)
            year_soup = BeautifulSoup(decode_html(year_content), 'html.parser')
            overview_url = None
            for a in year_soup.find_all('a'):
                text = a.get_text(' ', strip=True)
                href = a.get('href')
                if href and '货币统计概览' in text:
                    overview_url = urljoin(year_final, href)
                    break
            if not overview_url:
                errors.append(f'{year_url}: 未发现货币统计概览入口')
                continue
            overview_content, overview_final = fetch(overview_url)
            overview_soup = BeautifulSoup(decode_html(overview_content), 'html.parser')
            candidates = []
            for a in overview_soup.find_all('a'):
                href = a.get('href') or ''
                text = a.get_text(' ', strip=True).lower()
                if href.lower().endswith(('.xlsx', '.xls')) or text in ('xls', 'xlsx'):
                    candidates.append(urljoin(overview_final, href))
            for candidate in candidates:
                if not candidate.lower().endswith('.xlsx'):
                    continue
                try:
                    data, resolved = fetch(candidate)
                    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
                    first = normalize_label(wb.active['A1'].value)
                    if '货币当局资产负债表' in first:
                        return {
                            'source_url': resolved,
                            'source_title': '货币当局资产负债表',
                            'published_date': source_date_from_url(resolved),
                            'content': data,
                            'errors': errors,
                        }
                except Exception as exc:
                    errors.append(f'{candidate}: {exc}')
        except Exception as exc:
            errors.append(f'{year_url}: {exc}')
    raise RuntimeError('未找到可解析的货币当局资产负债表 xlsx；' + '; '.join(errors[:8]))


def parse_balance_sheet_xlsx(content, source_meta):
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    periods_by_col = {}
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        for cell in row:
            period = period_from_header(cell.value)
            if period:
                periods_by_col[cell.column] = period
        if periods_by_col:
            break
    if not periods_by_col:
        raise RuntimeError('资产负债表未发现月度列头')

    values = {period: {} for period in periods_by_col.values()}
    for row in ws.iter_rows(values_only=False):
        label = normalize_label(row[0].value if row else '')
        if not label:
            continue
        matched_code = None
        for code, needles in ROW_MATCHERS:
            if any(needle in label for needle in needles):
                matched_code = code
                break
        if not matched_code:
            continue
        for cell in row[1:]:
            period = periods_by_col.get(cell.column)
            if period and isinstance(cell.value, (int, float)):
                values[period][matched_code] = float(cell.value)

    observations = []
    for period, row_values in values.items():
        if not row_values:
            continue
        for code, name in BASE_INDICATORS.items():
            value = row_values.get(code)
            if value is None:
                continue
            observations.append({
                'indicator_code': code,
                'indicator_name': name,
                'period': period,
                'frequency': 'monthly',
                'value': value,
                'unit': '亿元',
                'data_status': 'official',
                'formula': None,
                **source_meta,
            })
        for code, (formula, numerator, denominator) in DERIVED_INDICATORS.items():
            num = row_values.get(numerator)
            den = row_values.get(denominator)
            if num is None or not den:
                continue
            observations.append({
                'indicator_code': code,
                'indicator_name': formula,
                'period': period,
                'frequency': 'monthly',
                'value': num / den * 100.0,
                'unit': '%',
                'data_status': 'derived',
                'formula': f'{numerator} / {denominator} * 100',
                **source_meta,
            })
    return observations


def update_pboc_balance_sheet(db_path):
    started = datetime.now().isoformat()
    parser_errors = []
    with connect(db_path) as conn:
        ensure_pboc_balance_sheet_tables(conn)
        try:
            attachment = discover_balance_sheet_attachment()
            parser_errors.extend(attachment.get('errors') or [])
            source_meta = {
                'source_name': PBOC_SOURCE_NAME,
                'source_type': PBOC_BALANCE_SOURCE_TYPE,
                'source_url': attachment['source_url'],
                'source_title': attachment['source_title'],
                'published_date': attachment.get('published_date'),
                'parser_notes': '解析自中国人民银行“货币统计概览”下的“货币当局资产负债表”xlsx 附件；占比字段按表内项目除以总资产计算。',
            }
            observations = parse_balance_sheet_xlsx(attachment['content'], source_meta)
            if not observations:
                raise RuntimeError('资产负债表附件未解析出目标指标')
            now = datetime.now().isoformat()
            for obs in observations:
                conn.execute('''INSERT INTO pboc_balance_sheet_observations (
                    indicator_code,indicator_name,period,frequency,value,unit,data_status,
                    source_name,source_type,source_url,source_title,published_date,parser_notes,formula,updated_at
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(indicator_code,period,source_url) DO UPDATE SET
                    value=excluded.value, data_status=excluded.data_status, parser_notes=excluded.parser_notes,
                    formula=excluded.formula, updated_at=excluded.updated_at
                ''', (
                    obs['indicator_code'], obs['indicator_name'], obs['period'], obs['frequency'], obs['value'], obs['unit'],
                    obs['data_status'], obs['source_name'], obs['source_type'], obs['source_url'], obs['source_title'],
                    obs['published_date'], obs['parser_notes'], obs['formula'], now
                ))
            conn.execute('''INSERT INTO fiscal_debt_sources (
                source_name,source_type,source_url,source_title,published_date,parser_notes,raw_text,
                parsed_indicators,status,error,updated_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(source_url) DO UPDATE SET
                source_title=excluded.source_title, published_date=excluded.published_date,
                parser_notes=excluded.parser_notes, raw_text=excluded.raw_text,
                parsed_indicators=excluded.parsed_indicators, status=excluded.status,
                error=excluded.error, updated_at=excluded.updated_at
            ''', (
                PBOC_SOURCE_NAME, PBOC_BALANCE_SOURCE_TYPE, source_meta['source_url'], source_meta['source_title'],
                source_meta['published_date'], source_meta['parser_notes'], 'xlsx attachment',
                json.dumps(sorted({o['indicator_code'] for o in observations}), ensure_ascii=False),
                'success', None, now
            ))
            conn.commit()
            return {
                'success': True,
                'started_at': started,
                'finished_at': now,
                'records': len(observations),
                'latest_period': max(o['period'] for o in observations),
                'source_url': source_meta['source_url'],
                'parser_errors': parser_errors,
            }
        except Exception as exc:
            now = datetime.now().isoformat()
            conn.execute('''INSERT OR REPLACE INTO fiscal_debt_sources (
                source_name,source_type,source_url,source_title,published_date,parser_notes,raw_text,
                parsed_indicators,status,error,updated_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?)''', (
                PBOC_SOURCE_NAME, PBOC_BALANCE_SOURCE_TYPE, PBOC_YEAR_INDEX, '货币当局资产负债表',
                None, '发现或解析货币当局资产负债表失败。', '',
                '[]', 'error', str(exc), now
            ))
            conn.commit()
            return {'success': False, 'error': str(exc), 'parser_errors': parser_errors}


def _coverage(conn):
    row = conn.execute('''SELECT count(*) records, min(period) earliest_period, max(period) latest_period,
        count(distinct period) months,
        sum(case when source_url is null or source_url='' then 1 else 0 end) missing_source_url
        FROM pboc_balance_sheet_observations''').fetchone()
    return dict(row)


def build_pboc_balance_sheet_payload(db_path):
    with connect(db_path) as conn:
        ensure_pboc_balance_sheet_tables(conn)
        rows = [dict(r) for r in conn.execute(
            'SELECT * FROM pboc_balance_sheet_observations ORDER BY period, indicator_code'
        ).fetchall()]
        coverage = _coverage(conn)
        sources = [dict(r) for r in conn.execute(
            'SELECT * FROM fiscal_debt_sources WHERE source_type=? ORDER BY updated_at DESC LIMIT 20',
            (PBOC_BALANCE_SOURCE_TYPE,)
        ).fetchall()]
    by_period = {}
    for row in rows:
        rec = by_period.setdefault(row['period'], {
            'period': row['period'],
            'source_name': row['source_name'],
            'source_type': row['source_type'],
            'source_url': row['source_url'],
            'source_title': row['source_title'],
            'published_date': row['published_date'],
            'parser_notes': row['parser_notes'],
            'data_status': 'official',
        })
        rec[row['indicator_code']] = row['value']
        rec[f"{row['indicator_code']}__status"] = row['data_status']
        if row.get('formula'):
            rec[f"{row['indicator_code']}__formula"] = row['formula']
    records = [by_period[p] for p in sorted(by_period)]
    latest = records[-1] if records else None
    return {
        'data_status': 'official' if records else 'missing',
        'latest_period': latest['period'] if latest else None,
        'records': records,
        'latest': latest,
        'coverage': coverage,
        'source_records': sources,
        'warnings': [] if records else ['央行资产负债表尚未成功接入；未生成 mock 数据。'],
        'notes': [
            '“对政府债权”是央行资产负债表内政府债权口径，不等于财政部国债余额，也不等于地方政府债务余额。',
            '央行资产负债表未单列地方政府债；地方债在本模块显示为 not_available，不做推断。',
        ],
        'local_government_debt_in_pboc_balance_sheet': {
            'data_status': 'not_available',
            'warning': '央行资产负债表没有地方政府债单列项目，不能推断为官方事实。',
        },
    }


def build_pboc_balance_sheet_debug(conn):
    ensure_pboc_balance_sheet_tables(conn)
    latest = [dict(r) for r in conn.execute('''
        SELECT * FROM pboc_balance_sheet_observations o
        WHERE period=(SELECT max(period) FROM pboc_balance_sheet_observations WHERE indicator_code=o.indicator_code)
        ORDER BY indicator_code
    ''').fetchall()]
    by_indicator = [dict(r) for r in conn.execute('''
        SELECT indicator_code, count(*) count, min(period) earliest_period, max(period) latest_period,
               group_concat(distinct data_status) data_statuses
        FROM pboc_balance_sheet_observations GROUP BY indicator_code ORDER BY indicator_code
    ''').fetchall()]
    status_dist = [dict(r) for r in conn.execute('''
        SELECT indicator_code, data_status, count(*) count
        FROM pboc_balance_sheet_observations GROUP BY indicator_code, data_status
        ORDER BY indicator_code, data_status
    ''').fetchall()]
    return {
        'coverage': _coverage(conn),
        'latest_by_indicator': latest,
        'by_indicator': by_indicator,
        'status_distribution': status_dist,
        'missing_source_url': conn.execute(
            "SELECT count(*) FROM pboc_balance_sheet_observations WHERE source_url IS NULL OR source_url=''"
        ).fetchone()[0],
    }
